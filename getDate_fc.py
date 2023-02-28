
import torch
from torch import nn
import numpy as np
import time as t
from torch.autograd import Variable
import random
from openpyxl import load_workbook

class Fclayer(nn.Module):
    def __init__(self, in_s, out_s, i):
        super().__init__()
        self.fc = nn.Sequential(
            nn.Linear(in_s, out_s),
            nn.ReLU()
        )
            
    def forward(self, x):
        t1 = t.time()
        x = self.fc(x)
        t2 = t.time()
        sheet.cell(i+1,1,t2-t1)
        return x  

i_max = 5000
wb = load_workbook('data_fc.xlsx')
sheet = wb.active

for i in range(i_max):
    in_s = random.randint(1,5000)
    out_s = random.randint(1,5000)
    fc = Fclayer(in_s,out_s,i)
    input_data = Variable(torch.rand(1,1,in_s))
    output = fc(input_data)
    # print(output.size()[2])
    sheet.cell(i+1,2,in_s)
    sheet.cell(i+1,3,out_s)
    print(i)

wb.save('data_fc.xlsx')
print("ending")