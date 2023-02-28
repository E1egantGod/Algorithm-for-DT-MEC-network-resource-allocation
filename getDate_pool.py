import torch
from torch import nn
import numpy as np
import time as t
from torch.autograd import Variable
import random
from openpyxl import load_workbook

class Poollayer(nn.Module):
    def __init__(self, kernel, i):
        super().__init__()
        self.pool = nn.Sequential(
            nn.MaxPool2d(kernel,1)
        )
            
    def forward(self, x):
        t1 = t.time()
        x = self.pool(x)
        t2 = t.time()
        sheet.cell(i+1,1,t2-t1)
        return x  

i_max = 5000
wb = load_workbook('data_pool.xlsx')
sheet = wb.active

for i in range(i_max):
    ch = random.randint(1,512)
    kernel_size = random.randint(1,8)
    stride = random.randint(1,4)
    in_size = random.randint(10,227)
    pool = Poollayer(kernel_size,i)
    input_data = Variable(torch.rand(1,ch,in_size,in_size))
    output = pool(input_data)
    # print(output.size()[2])
    sheet.cell(i+1,2,ch)
    sheet.cell(i+1,3,kernel_size)
    sheet.cell(i+1,4,output.size()[2])
    sheet.cell(i+1,5,stride)
    print(i)

wb.save('data_pool.xlsx')
print("ending")